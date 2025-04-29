import json
import os
import csv
import matplotlib.pyplot as plt
import numpy as np
import matplotlib
from openpyxl import Workbook  # Import for creating Excel files

matplotlib.use('Agg')  # Use the non-interactive Agg backend







class Results:
    def __init__(self):
        self.Curves = []  # List to store curve data
        self.PerformanceMetrics = {}  # Dictionary to store performance metrics
        self.Config = {}  # Dictionary to store configuration parameters
        self.Metrics = {}  # Dictionary to store additional metrics

    def add_curve(self, x_data, y_data, x_label="X", y_label="Y", name=None, plot_avg=False, plot_std=False, plotType="line"):
        """
        Add a curve to the results.
        :param x_data: List of X-axis data.
        :param y_data: List of Y-axis data.
        :param x_label: Label for the X-axis.
        :param y_label: Label for the Y-axis.
        :param name: Optional name for the curve.
        """
        curve = {
            "X": x_data,
            "Y": y_data,
            "XLabel": x_label,
            "YLabel": y_label,
            "Name": name,
            "EnableAvg": plot_avg,
            "EnableStd": plot_std,
            "PlotType": plotType
        }
        if plot_avg:           
            if plotType == "scatter":
                curve["Avg"] = np.mean(np.column_stack((x_data, y_data)), axis=0)
            else:
                curve["Avg"] = np.mean(y_data, axis=0)
        if plot_std:
            if plotType == "scatter":
                curve["Std"] = np.std(np.column_stack((x_data, y_data)), axis=0)
            else:
                curve["Std"] = np.std(y_data, axis=0)

        self.Curves.append(curve)
    def add_metric(self, name, value):
        """
        Add a performance metric.
        :param name: Name of the metric.
        :param value: Value of the metric.
        """
        self.Metrics[name] = value
    def set_performance(self, value):
        """
        Add a performance metric.
        :param name: Name of the metric.
        :param value: Value of the metric.
        """
        self.PerformanceMetrics = value

    def set_config(self, config):
        """
        Set the configuration parameters.
        :param config: Dictionary of configuration parameters.
        """
        self.Config = config

    def plot_curve(self, curve, path="./", dpi=300, bbox_inches='tight', format='png', overwrite=True):
        """
        Plot a curve and save it to a file.
        :param curve: The curve data to plot.
        :param path: The file path where the figure will be saved (default: './').
        :param dpi: The resolution of the saved figure in dots per inch.
        :param bbox_inches: Bounding box option for the figure (default: 'tight').
        :param format: The format to save the figure (e.g., 'png', 'pdf').
        :param overwrite: Whether to overwrite the file if it already exists (default: True).
        """
        fig, ax = plt.subplots()
        x_data = curve["X"]
        y_data = curve["Y"]
        x_label = curve["XLabel"]
        y_label = curve["YLabel"]
        name = curve["Name"]

        if curve["PlotType"] == "line":            
            if curve["EnableAvg"]:
                ax.plot(x_data, curve["Avg"], label="Avg", linestyle='--')
                if curve["EnableStd"]:
                     ax.fill_between(x_data, curve["Avg"] - curve["Std"], curve["Avg"] + curve["Std"], alpha=0.2, label="Std")
            if not curve["EnableAvg"] and not curve["EnableStd"]:
                ax.plot(x_data, y_data, label=name)
        elif curve["PlotType"] == "scatter":
            ax.scatter(x_data, y_data, label=name, alpha=0.6, color='blue')
            ax.errorbar(curve["Avg"][0], curve["Avg"][1], xerr=curve["Std"][0], yerr=curve["Std"][1], fmt='o', color='red', label='Avg ± Std', capsize=5)

        ax.set_xlabel(x_label)
        ax.set_ylabel(y_label)
        ax.set_title(name if name else "Curve")
        ax.grid(True)
        ax.legend()
        fig_path = os.path.join(path, name + "." + format)
        self.save_figure(fig, fig_path, dpi=dpi, bbox_inches=bbox_inches, format=format, overwrite=overwrite)
    
    def save_figure(self, fig, path, dpi=300, bbox_inches='tight', format='png', overwrite=True):
        """
        Save a matplotlib figure to a file.

        Parameters:
        - fig (matplotlib.figure.Figure): The figure to save.
        - path (str): The file path where the figure will be saved.
        - dpi (int): The resolution of the saved figure in dots per inch.
        - bbox_inches (str): Bounding box option for the figure (default: 'tight').
        - format (str): The format to save the figure (e.g., 'png', 'pdf').
        - overwrite (bool): Whether to overwrite the file if it already exists (default: True).

        Returns:
        - bool: True if the file was saved successfully, False otherwise.
        """
        try:
            if not overwrite and os.path.exists(path):
                print(f"File already exists and overwrite is disabled: {path}")
                return False

            self.ensure_directory_exists(path)
            fig.savefig(path, dpi=dpi, bbox_inches=bbox_inches, format=format)
            plt.close(fig)
            print(f"Figure saved successfully: {path}")
            return True
        except Exception as e:
            print(f"Failed to save figure: {e}")
            return False
    def ensure_directory_exists(self, path):
        """
        Ensure that the directory for the given path exists.

        Parameters:
        - path (str): The file path where the directory should exist.
        """
        os.makedirs(os.path.dirname(path), exist_ok=True)

    def save_results(self, path, overwrite=True):
        self.ensure_directory_exists(path)
        self.result_directory = path

        # Save plots
        for curve in self.Curves:
            self.plot_curve(curve=curve, path=path, overwrite=overwrite)

        performance_metrics_path = os.path.join(path, "performance_metrics.json")
        config_path = os.path.join(path, "config.json")

        if not overwrite and (os.path.exists(performance_metrics_path) or os.path.exists(config_path)):
            print("Files already exist and overwrite is disabled.")
            return False

        try:
            with open(performance_metrics_path, 'w', encoding='utf-8') as perf_file:
                json.dump(self.PerformanceMetrics, perf_file, indent=4)
                print(f"Performance metrics saved successfully: {performance_metrics_path}")

            with open(config_path, 'w', encoding='utf-8') as config_file:
                json.dump(self.Config, config_file, indent=4)
                print(f"Config saved successfully: {config_path}")
        except Exception as e:
            print(f"Failed to save JSON files: {e}")
            return False

        # Save metrics
        metrics_path = os.path.join(path, "metrics.csv")
        headers = ["Metric", "Value"]
        data = [[key, value] for key, value in self.Metrics.items()]
        self.save_csv(data, headers, metrics_path, overwrite=overwrite)

        # Save all curves in a single Excel file
        excel_path = os.path.join(path, "curves.xlsx")
        self.save_curves_to_excel(excel_path, overwrite=overwrite)

    def save_curves_to_excel(self, path, overwrite=True):
        """
        Save all curves to a single Excel file, with each curve in a separate sheet.

        Parameters:
        - path (str): The file path where the Excel file will be saved.
        - overwrite (bool): Whether to overwrite the file if it already exists (default: True).
        """
        if not overwrite and os.path.exists(path):
            print(f"File already exists and overwrite is disabled: {path}")
            return False

        try:
            workbook = Workbook()
            for curve in self.Curves:
                sheet_name = curve["Name"] if curve["Name"] else "Unnamed_Curve"
                sheet_name = sheet_name[:31]  # Excel sheet names are limited to 31 characters
                sheet = workbook.create_sheet(title=sheet_name)

                x_data = curve["X"]
                y_data = curve["Y"]

                #TO-DO: Consertar para optimal points
                # Handle multiple Y datasets (e.g., y_data is a list of lists)
                if isinstance(y_data[0], list) or isinstance(y_data[0], np.ndarray):
                    headers = ["X"] + [f"Y_{i+1}" for i in range(len(y_data))]
                    if curve["EnableAvg"]:
                        headers.append("Avg")
                    if curve["EnableStd"]:
                        headers.append("Std")

                    sheet.append(headers)
                    for i in range(len(x_data)):
                        row = [x_data[i]] + [y[i] if i < len(y) else "" for y in y_data]
                        if curve["EnableAvg"]:
                            row.append(curve["Avg"][i])
                        if curve["EnableStd"]:
                            row.append(curve["Std"][i])
                        sheet.append(row)
                else:
                    headers = ["X", "Y"]
                    sheet.append(headers)
                    for i in range(len(x_data)):
                        sheet.append([x_data[i], y_data[i]])

            # Remove the default sheet created by openpyxl
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

            workbook.save(path)
            print(f"Excel file saved successfully: {path}")
            return True
        except Exception as e:
            print(f"Failed to save Excel file: {e}")
            return False

    def save_csv(self, data, headers, path, overwrite=True):
        """
        Save data to a CSV file.

        Parameters:
        - data (list of lists): The rows of data to save.
        - headers (list): The column headers for the CSV file.
        - path (str): The file path where the CSV will be saved.
        - overwrite (bool): Whether to overwrite the file if it already exists (default: True).

        Returns:
        - bool: True if the file was saved successfully, False otherwise.
        """
        try:
            if not overwrite and os.path.exists(path):
                print(f"File already exists and overwrite is disabled: {path}")
                return False

            self.ensure_directory_exists(path)
            with open(path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                writer.writerows(data)
            print(f"CSV saved successfully: {path}")
            return True
        except Exception as e:
            print(f"Failed to save CSV: {e}")
            return False